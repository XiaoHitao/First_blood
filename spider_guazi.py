
#coding:utf-8 编码模式为UTF8
'''A spider for guazi Data'''

import re 
import random
import time
import requests
import lxml
import lxml.etree
import pandas as pd
import openpyxl
from openpyxl import load_workbook



def get_guazi_data(header,car_id,sample_id):
	'''get car location and fetch the car url'''
	
	# get redict url throuht car_ID and location
	car_url = "https://www.guazi.com/bj/buy/_{}/?input=1".format(car_id)
	lis = list()
	try:  
		data = requests.get(car_url,headers = header,allow_redirects=False)	
		location = data.headers['location']		# get the real url of the car 
		new_url = 'https://www.guazi.com' + location  
		data_needed = requests.get(new_url,headers = header)
		
		#  check the spider statue through status code
		#  if (data_needed.status_code == 200):
		#  get current city\brand name\current car id\count shortcome 
		mytree = lxml.etree.HTML(data_needed.content.decode('utf-8'))
		cur_city= re.sub('[ \n\r]','',mytree.xpath('//*[@id="jstop"]/div/div[1]/p/text()')[0]) 
		lis.append(cur_city)
		brand_name= re.sub('[ \n\r]','',mytree.xpath('/html/body/div[5]/div[3]/div[2]/h2/text()')[0]) 
		lis.append(brand_name)
		cur_id = mytree.xpath('/html/body/div[5]/div[2]/div[2]/text()')[0].split('-')[-1] 
		lis.append(cur_id)
		span_list = mytree.xpath("//span[contains(@class ,'fc-org-text')]")  
		error_con = 0 
		for span in span_list:
			error = re.match(r'^[0-9]+',span.xpath('./text()')[0]).group(0)
			error_con = int(error) + error_con
		lis.append(error_con)		
		picture_url = mytree.xpath('//*[@id="page-slide"]/div[1]/ul/li[1]/img/@data-src')[0] #  //*[@id="page-slide"]/div[1]/ul/li[1]
		# print picture_url

		picture = requests.get(picture_url,headers = header)
		car_id = str(car_id)
		picture_name= 'C:\\Users\\peng\\Desktop\\' + car_id + '.jpg'
		with open(picture_name,'wb') as fid:
			fid.write(picture.content)
		try:
			sold = mytree.xpath('//*[@id="page-slide"]/div[1]/div[2]/@class')
			if sold[0] == 'sold-icon-yishou':
				lis.append('sold')
		except Exception as e:
			pass
		
		return lis

	
	except:
		print 'car ID {} is not found'.format(car_id)
	  	lis.append('error id')
	  	lis.append(car_id)
	  	return lis
	


def Load_Car_Id():
	""" Load car id from excel"""

 	excel_path = 'C:\\Users\\peng\\Desktop\\c.xlsx'
 	d = pd.read_excel(excel_path, sheetname=None)
 	car_list = d['Sheet1'].carid
 	return car_list


def write_to_excel(file_path,iteams):
	wb = load_workbook(file_path)
	wb1 = wb.active
	i = 2 
	for iteam in iteams:
		j = 2	
		for inf in iteam:
			wb1.cell(i,j,inf)
 			# print inf
	 		j+=1
	 	i+=1

	wb.save(file_path)


def Ramdon_id_Create(car_list,Ramdon_Num):
	""" Sample from all cars"""
	samples = list()
	Random_list = random.sample(range(0,len(car_list)),Ramdon_Num)
	for num in Random_list:
		samples.append(car_list[num])
	return samples


def main():
	headers = {
		   # 'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
           # 'Accept-Encoding':'gzip, deflate,br',
           # 'Accept-Language':'zh-CN,zh;q=0.9',
           # 'Cache-Control':'no-cache',
           # 'Connection':'keep-alive',
           # 'cookie':'cityDomain=tj; cainfo=%7B%22ca_s%22%3A%22sem_360ss%22%2C%22ca_n%22%3A%22360pc_shouye%22%2C%22ca_i%22%3A%22-%22%2C%22ca_medium%22%3A%22-%22%2C%22ca_term%22%3A%22%7Bkeyword%7D%22%2C%22ca_content%22%3A%22-%22%2C%22ca_campaign%22%3A%22-%22%2C%22ca_kw%22%3A%22%25e7%2593%259c%25e5%25ad%2590%25e4%25ba%258c%25e6%2589%258b%25e8%25bd%25a6%25e7%259b%25b4%25e4%25b9%25b0%25e5%25ae%2598%25e7%25bd%2591%22%2C%22keyword%22%3A%22-%22%2C%22ca_keywordid%22%3A%2211680840494%22%2C%22scode%22%3A%2210103213212%22%2C%22ca_transid%22%3Anull%2C%22platform%22%3A%221%22%2C%22version%22%3A1%2C%22ca_b%22%3A%22-%22%2C%22ca_a%22%3A%22-%22%2C%22display_finance_flag%22%3A%22-%22%2C%22client_ab%22%3A%22-%22%2C%22guid%22%3A%225c4b7da5-3581-4aab-a484-9aae3512e621%22%2C%22sessionid%22%3A%221cd7d36a-4deb-4f94-eb58-7a9995f63871%22%7D; clueSourceCode=10103213212%2300; uuid=5c4b7da5-3581-4aab-a484-9aae3512e621; preTime=%7B%22last%22%3A1542631338%2C%22this%22%3A1542631222%2C%22pre%22%3A1542631222%7D; ganji_uuid=8822428375354962391098; sessionid=1cd7d36a-4deb-4f94-eb58-7a9995f63871; lg=1; close_finance_popup=2018-11-19; antipas=2e82rOmZ819210317098429715857',
           # 'Host':'www.guazi.com',
           # 'Pragma':'no-cache',
           # 'Referer':'https://www.guazi.com/sz/buy/o1r3_16_6/',
           # 'Upgrade-Insecure-Requests':'1',
           # 'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko'}
           'cookie':'uuid=fb667ef6-d8b6-4b05-88d8-52de8a4da966; ganji_uuid=2515070520096047758509; lg=1; clueSourceCode=10103000312%2300; antipas=9s8716JvRv746L4109250952232; sessionid=d28af670-cf8d-45c3-84bd-2cc9246e2718; close_finance_popup=2018-11-19; _gl_tracker=%7B%22ca_source%22%3A%22-%22%2C%22ca_name%22%3A%22-%22%2C%22ca_kw%22%3A%22-%22%2C%22ca_id%22%3A%22-%22%2C%22ca_s%22%3A%22self%22%2C%22ca_n%22%3A%22-%22%2C%22ca_i%22%3A%22-%22%2C%22sid%22%3A22084896923%7D; cityDomain=bj; Hm_lvt_936a6d5df3f3d309bda39e92da3dd52f=1542636723; cainfo=%7B%22ca_s%22%3A%22seo_baidu%22%2C%22ca_n%22%3A%22tbmkbturl%22%2C%22ca_i%22%3A%22-%22%2C%22ca_medium%22%3A%22-%22%2C%22ca_term%22%3A%22-%22%2C%22ca_content%22%3A%22-%22%2C%22ca_campaign%22%3A%22-%22%2C%22ca_kw%22%3A%22-%22%2C%22keyword%22%3A%22-%22%2C%22ca_keywordid%22%3A%22-%22%2C%22scode%22%3A%2210103000312%22%2C%22ca_transid%22%3Anull%2C%22platform%22%3A%221%22%2C%22version%22%3A1%2C%22ca_b%22%3A%22-%22%2C%22ca_a%22%3A%22-%22%2C%22display_finance_flag%22%3A%22-%22%2C%22client_ab%22%3A%22-%22%2C%22guid%22%3A%22fb667ef6-d8b6-4b05-88d8-52de8a4da966%22%2C%22sessionid%22%3A%22d28af670-cf8d-45c3-84bd-2cc9246e2718%22%7D; preTime=%7B%22last%22%3A1542638931%2C%22this%22%3A1542447546%2C%22pre%22%3A1542447546%7D; Hm_lpvt_936a6d5df3f3d309bda39e92da3dd52f=1542638934',
           'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'}

	car_list = Load_Car_Id()
	print 'The car amount is {}'.format(len(car_list))
	# Sample_Num = input('key in the number of the car : ')
	Sample_Num = 20
	sample_id = Ramdon_id_Create(car_list,Sample_Num)
	print '---------SAMPLE DONE--------------'
	car_data = list() 
	for car_id in sample_id:
		N_done = (1+sample_id.index(car_id))*100/len(sample_id)
		# print (sample_id.index(car_id))
		print('Downloading Car '+ str(car_id) +'>'*(int(N_done * 0.2)) + '.'*(20 - int(N_done*0.20)) + '[%d %%]'%(N_done))
		car_data.append(get_guazi_data(headers,car_id,sample_id))
		t = random.uniform(0,2)
		time.sleep(t)
	# print car_data
	write_to_excel('C:\\Users\\peng\\Desktop\\v.xlsx',car_data)


if __name__ == '__main__':
	main()